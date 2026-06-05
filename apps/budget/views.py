from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    TemplateView,
    UpdateView,
)

from apps.core.permissions import accessible_projects, can_manage_project

from .forms import (
    BoQItemForm,
    CostCodeForm,
    CostEntryForm,
    SubcontractBackChargeForm,
    SubcontractClaimForm,
    SubcontractForm,
    SubcontractPerformanceReviewForm,
)
from .models import (
    BoQItem,
    CostCode,
    CostEntry,
    Subcontract,
    SubcontractBackCharge,
    SubcontractClaim,
    SubcontractPerformanceReview,
)


# ---------------------------------------------------------------------------
# Mixins
# ---------------------------------------------------------------------------


class ProjectMixin(LoginRequiredMixin):
    """Injects the current project (from URL kwarg ``project_pk``) into context."""

    def get_project(self):
        if not hasattr(self, "_project"):
            self._project = get_object_or_404(
                accessible_projects(self.request.user),
                pk=self.kwargs["project_pk"],
            )
        return self._project

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["project"] = self.get_project()
        return ctx


# ---------------------------------------------------------------------------
# Budget Dashboard
# ---------------------------------------------------------------------------


class BudgetDashboardView(ProjectMixin, TemplateView):
    """High-level budget vs actual table with RAG status per cost code."""

    template_name = "budget/dashboard.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        project = self.get_project()
        cost_codes = CostCode.objects.filter(project=project).order_by("code")

        rows = []
        total_budget = Decimal("0.00")
        total_control_budget = Decimal("0.00")
        total_committed = Decimal("0.00")
        total_actual = Decimal("0.00")
        total_etc = Decimal("0.00")
        total_efc = Decimal("0.00")
        total_provisional_sum = Decimal("0.00")
        total_contingency_drawdown = Decimal("0.00")

        for cc in cost_codes:
            committed = cc.total_committed
            actual = cc.total_actual
            spent = cc.total_spent
            efc = cc.estimate_at_completion
            total_budget += cc.budget_amount
            total_control_budget += cc.control_budget
            total_committed += committed
            total_actual += actual
            total_etc += cc.forecast_etc
            total_efc += efc
            total_provisional_sum += cc.provisional_sum
            total_contingency_drawdown += cc.contingency_drawdown
            rows.append(
                {
                    "cost_code": cc,
                    "budget": cc.budget_amount,
                    "control_budget": cc.control_budget,
                    "committed": committed,
                    "actual": actual,
                    "etc": cc.forecast_etc,
                    "efc": efc,
                    "spent": spent,
                    "variance": cc.variance,
                    "variance_pct": cc.variance_percentage,
                    "forecast_variance": cc.forecast_variance,
                    "forecast_pct": cc.forecast_variance_percentage,
                    "rag": cc.rag_status,
                    "forecast_rag": cc.forecast_rag_status,
                }
            )

        total_spent = total_committed + total_actual
        contract_value = project.contract_value
        forecast_margin = contract_value - total_efc
        forecast_margin_pct = (
            (forecast_margin / contract_value * 100).quantize(Decimal("0.01"))
            if contract_value
            else Decimal("0.00")
        )
        ctx.update(
            {
                "rows": rows,
                "contract_value": contract_value,
                "total_budget": total_budget,
                "total_control_budget": total_control_budget,
                "total_committed": total_committed,
                "total_actual": total_actual,
                "total_etc": total_etc,
                "total_efc": total_efc,
                "total_spent": total_spent,
                "total_variance": total_control_budget - total_spent,
                "total_forecast_variance": total_control_budget - total_efc,
                "total_provisional_sum": total_provisional_sum,
                "total_contingency_drawdown": total_contingency_drawdown,
                "forecast_margin": forecast_margin,
                "forecast_margin_pct": forecast_margin_pct,
            }
        )
        return ctx


# ---------------------------------------------------------------------------
# Cost Code views
# ---------------------------------------------------------------------------


class CostCodeListView(ProjectMixin, ListView):
    model = CostCode
    template_name = "budget/costcode_list.html"
    context_object_name = "cost_codes"

    def get_queryset(self):
        return CostCode.objects.filter(project=self.get_project()).order_by("code")


class CostCodeCreateView(ProjectMixin, CreateView):
    model = CostCode
    form_class = CostCodeForm
    template_name = "budget/costcode_form.html"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["project"] = self.get_project()
        return kwargs

    def form_valid(self, form):
        if not can_manage_project(self.request.user, self.get_project()):
            messages.error(self.request, "You do not have permission to create cost codes for this project.")
            return redirect(reverse_lazy("budget:costcode-list", kwargs={"project_pk": self.get_project().pk}))
        form.instance.project = self.get_project()
        form.instance.created_by = self.request.user
        messages.success(self.request, "Cost code created successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("budget:costcode-list", kwargs={"project_pk": self.get_project().pk})


class CostCodeUpdateView(ProjectMixin, UpdateView):
    model = CostCode
    form_class = CostCodeForm
    template_name = "budget/costcode_form.html"

    def get_queryset(self):
        return CostCode.objects.filter(project=self.get_project())

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["project"] = self.get_project()
        return kwargs

    def form_valid(self, form):
        if not can_manage_project(self.request.user, self.get_project()):
            messages.error(self.request, "You do not have permission to update cost codes for this project.")
            return redirect(reverse_lazy("budget:costcode-list", kwargs={"project_pk": self.get_project().pk}))
        form.instance.updated_by = self.request.user
        messages.success(self.request, "Cost code updated successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("budget:costcode-list", kwargs={"project_pk": self.get_project().pk})


# ---------------------------------------------------------------------------
# BoQ views
# ---------------------------------------------------------------------------


class BoQListView(ProjectMixin, ListView):
    model = BoQItem
    template_name = "budget/boq_list.html"
    context_object_name = "boq_items"

    def get_queryset(self):
        qs = BoQItem.objects.filter(project=self.get_project()).select_related("cost_code")
        trade = self.request.GET.get("trade")
        if trade:
            qs = qs.filter(trade_section=trade)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["trade_choices"] = CostCode.CATEGORY_CHOICES
        ctx["selected_trade"] = self.request.GET.get("trade", "")
        total = sum(item.amount for item in ctx["boq_items"])
        ctx["total_amount"] = total
        return ctx


class BoQItemCreateView(ProjectMixin, CreateView):
    model = BoQItem
    form_class = BoQItemForm
    template_name = "budget/boq_form.html"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["project"] = self.get_project()
        return kwargs

    def form_valid(self, form):
        if not can_manage_project(self.request.user, self.get_project()):
            messages.error(self.request, "You do not have permission to create BoQ items for this project.")
            return redirect(reverse_lazy("budget:boq-list", kwargs={"project_pk": self.get_project().pk}))
        form.instance.project = self.get_project()
        form.instance.created_by = self.request.user
        messages.success(self.request, "BoQ item added successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("budget:boq-list", kwargs={"project_pk": self.get_project().pk})


class BoQItemUpdateView(ProjectMixin, UpdateView):
    model = BoQItem
    form_class = BoQItemForm
    template_name = "budget/boq_form.html"

    def get_queryset(self):
        return BoQItem.objects.filter(project=self.get_project())

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["project"] = self.get_project()
        return kwargs

    def form_valid(self, form):
        if not can_manage_project(self.request.user, self.get_project()):
            messages.error(self.request, "You do not have permission to update BoQ items for this project.")
            return redirect(reverse_lazy("budget:boq-list", kwargs={"project_pk": self.get_project().pk}))
        form.instance.updated_by = self.request.user
        messages.success(self.request, "BoQ item updated successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("budget:boq-list", kwargs={"project_pk": self.get_project().pk})


# ---------------------------------------------------------------------------
# Cost Entry views
# ---------------------------------------------------------------------------


class CostEntryListView(ProjectMixin, ListView):
    model = CostEntry
    template_name = "budget/costentry_list.html"
    context_object_name = "cost_entries"
    paginate_by = 30

    def get_queryset(self):
        qs = CostEntry.objects.filter(project=self.get_project()).select_related("cost_code", "boq_item")
        entry_type = self.request.GET.get("type")
        cost_code = self.request.GET.get("cost_code")
        if entry_type:
            qs = qs.filter(entry_type=entry_type)
        if cost_code:
            qs = qs.filter(cost_code_id=cost_code)
        return qs.order_by("-date")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        project = self.get_project()
        ctx["cost_codes"] = CostCode.objects.filter(project=project)
        ctx["entry_type_choices"] = CostEntry.ENTRY_TYPE_CHOICES
        ctx["selected_type"] = self.request.GET.get("type", "")
        ctx["selected_cost_code"] = self.request.GET.get("cost_code", "")
        return ctx


class CostEntryCreateView(ProjectMixin, CreateView):
    model = CostEntry
    form_class = CostEntryForm
    template_name = "budget/costentry_form.html"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["project"] = self.get_project()
        return kwargs

    def form_valid(self, form):
        if not can_manage_project(self.request.user, self.get_project()):
            messages.error(self.request, "You do not have permission to record cost entries for this project.")
            return redirect(reverse_lazy("budget:costentry-list", kwargs={"project_pk": self.get_project().pk}))
        form.instance.project = self.get_project()
        form.instance.created_by = self.request.user
        messages.success(self.request, "Cost entry recorded successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("budget:costentry-list", kwargs={"project_pk": self.get_project().pk})


class CostEntryUpdateView(ProjectMixin, UpdateView):
    model = CostEntry
    form_class = CostEntryForm
    template_name = "budget/costentry_form.html"

    def get_queryset(self):
        return CostEntry.objects.filter(project=self.get_project())

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["project"] = self.get_project()
        return kwargs

    def form_valid(self, form):
        if not can_manage_project(self.request.user, self.get_project()):
            messages.error(self.request, "You do not have permission to update cost entries for this project.")
            return redirect(reverse_lazy("budget:costentry-list", kwargs={"project_pk": self.get_project().pk}))
        form.instance.updated_by = self.request.user
        messages.success(self.request, "Cost entry updated successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("budget:costentry-list", kwargs={"project_pk": self.get_project().pk})


# ---------------------------------------------------------------------------
# Subcontract views
# ---------------------------------------------------------------------------


class SubcontractListView(ProjectMixin, ListView):
    model = Subcontract
    template_name = "budget/subcontract_list.html"
    context_object_name = "subcontracts"

    def get_queryset(self):
        qs = Subcontract.objects.filter(project=self.get_project())
        status = self.request.GET.get("status")
        if status:
            qs = qs.filter(status=status)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["status_choices"] = Subcontract.STATUS_CHOICES
        ctx["selected_status"] = self.request.GET.get("status", "")
        return ctx


class SubcontractDetailView(ProjectMixin, DetailView):
    model = Subcontract
    template_name = "budget/subcontract_detail.html"
    context_object_name = "subcontract"

    def get_queryset(self):
        return Subcontract.objects.filter(project=self.get_project()).prefetch_related(
            "claims", "backcharges", "performance_reviews"
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        subcontract = self.object
        ctx["claims"] = subcontract.claims.order_by("-submitted_date", "-claim_number")
        ctx["backcharges"] = subcontract.backcharges.order_by("-date")
        ctx["performance_reviews"] = subcontract.performance_reviews.order_by("-review_date")
        return ctx


class SubcontractCreateView(ProjectMixin, CreateView):
    model = Subcontract
    form_class = SubcontractForm
    template_name = "budget/subcontract_form.html"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["project"] = self.get_project()
        return kwargs

    def form_valid(self, form):
        if not can_manage_project(self.request.user, self.get_project()):
            messages.error(self.request, "You do not have permission to create subcontracts for this project.")
            return redirect(reverse_lazy("budget:subcontract-list", kwargs={"project_pk": self.get_project().pk}))
        form.instance.project = self.get_project()
        form.instance.created_by = self.request.user
        messages.success(self.request, "Subcontract created successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("budget:subcontract-list", kwargs={"project_pk": self.get_project().pk})


class SubcontractUpdateView(ProjectMixin, UpdateView):
    model = Subcontract
    form_class = SubcontractForm
    template_name = "budget/subcontract_form.html"

    def get_queryset(self):
        return Subcontract.objects.filter(project=self.get_project())

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["project"] = self.get_project()
        return kwargs

    def form_valid(self, form):
        if not can_manage_project(self.request.user, self.get_project()):
            messages.error(self.request, "You do not have permission to update subcontracts for this project.")
            return redirect(reverse_lazy("budget:subcontract-list", kwargs={"project_pk": self.get_project().pk}))
        form.instance.updated_by = self.request.user
        messages.success(self.request, "Subcontract updated successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("budget:subcontract-list", kwargs={"project_pk": self.get_project().pk})


class SubcontractChildMixin(ProjectMixin):
    def get_subcontract(self):
        if not hasattr(self, "_subcontract"):
            self._subcontract = get_object_or_404(
                Subcontract,
                pk=self.kwargs["subcontract_pk"],
                project=self.get_project(),
            )
        return self._subcontract

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["subcontract"] = self.get_subcontract()
        return ctx

    def get_success_url(self):
        subcontract = self.get_subcontract()
        return reverse_lazy(
            "budget:subcontract-detail",
            kwargs={"project_pk": subcontract.project_id, "pk": subcontract.pk},
        )


class SubcontractClaimCreateView(SubcontractChildMixin, CreateView):
    model = SubcontractClaim
    form_class = SubcontractClaimForm
    template_name = "budget/subcontract_claim_form.html"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["subcontract"] = self.get_subcontract()
        return kwargs

    def form_valid(self, form):
        subcontract = self.get_subcontract()
        if not can_manage_project(self.request.user, subcontract.project):
            messages.error(self.request, "You do not have permission to create subcontract claims for this project.")
            return redirect(self.get_success_url())
        form.instance.subcontract = subcontract
        form.instance.created_by = self.request.user
        messages.success(self.request, "Subcontract claim recorded.")
        return super().form_valid(form)


class SubcontractClaimUpdateView(SubcontractChildMixin, UpdateView):
    model = SubcontractClaim
    form_class = SubcontractClaimForm
    template_name = "budget/subcontract_claim_form.html"

    def get_queryset(self):
        return SubcontractClaim.objects.filter(subcontract=self.get_subcontract())

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["subcontract"] = self.get_subcontract()
        return kwargs

    def form_valid(self, form):
        subcontract = self.get_subcontract()
        if not can_manage_project(self.request.user, subcontract.project):
            messages.error(self.request, "You do not have permission to update subcontract claims for this project.")
            return redirect(self.get_success_url())
        form.instance.updated_by = self.request.user
        messages.success(self.request, "Subcontract claim updated.")
        return super().form_valid(form)


class SubcontractBackChargeCreateView(SubcontractChildMixin, CreateView):
    model = SubcontractBackCharge
    form_class = SubcontractBackChargeForm
    template_name = "budget/subcontract_backcharge_form.html"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["subcontract"] = self.get_subcontract()
        return kwargs

    def form_valid(self, form):
        subcontract = self.get_subcontract()
        if not can_manage_project(self.request.user, subcontract.project):
            messages.error(self.request, "You do not have permission to create subcontract back-charges for this project.")
            return redirect(self.get_success_url())
        form.instance.subcontract = subcontract
        form.instance.created_by = self.request.user
        messages.success(self.request, "Subcontract back-charge recorded.")
        return super().form_valid(form)


class SubcontractBackChargeUpdateView(SubcontractChildMixin, UpdateView):
    model = SubcontractBackCharge
    form_class = SubcontractBackChargeForm
    template_name = "budget/subcontract_backcharge_form.html"

    def get_queryset(self):
        return SubcontractBackCharge.objects.filter(subcontract=self.get_subcontract())

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["subcontract"] = self.get_subcontract()
        return kwargs

    def form_valid(self, form):
        subcontract = self.get_subcontract()
        if not can_manage_project(self.request.user, subcontract.project):
            messages.error(self.request, "You do not have permission to update subcontract back-charges for this project.")
            return redirect(self.get_success_url())
        form.instance.updated_by = self.request.user
        messages.success(self.request, "Subcontract back-charge updated.")
        return super().form_valid(form)


class SubcontractPerformanceReviewCreateView(SubcontractChildMixin, CreateView):
    model = SubcontractPerformanceReview
    form_class = SubcontractPerformanceReviewForm
    template_name = "budget/subcontract_performance_form.html"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["subcontract"] = self.get_subcontract()
        kwargs["user"] = self.request.user
        return kwargs

    def form_valid(self, form):
        subcontract = self.get_subcontract()
        if not can_manage_project(self.request.user, subcontract.project):
            messages.error(self.request, "You do not have permission to review subcontractor performance for this project.")
            return redirect(self.get_success_url())
        form.instance.subcontract = subcontract
        form.instance.reviewer = self.request.user
        form.instance.created_by = self.request.user
        messages.success(self.request, "Subcontractor performance review recorded.")
        return super().form_valid(form)


# ---------------------------------------------------------------------------
# BoQ Import (CSV / Excel)
# ---------------------------------------------------------------------------


class BoQImportView(ProjectMixin, View):
    """
    Import BoQ items from a CSV or Excel file.

    Expected columns (case-insensitive, order flexible):
    item_number, description, unit, quantity, unit_rate, section (optional)
    """

    template_name = "budget/boq_import.html"

    def get(self, request, project_pk):
        from django.shortcuts import render
        project = self.get_project()
        return render(request, self.template_name, {"project": project, "breadcrumbs": [
            {"label": "Projects", "url": "/projects/"},
            {"label": project.name, "url": project.get_absolute_url()},
            {"label": "BoQ Import"},
        ]})

    def post(self, request, project_pk):
        from django.shortcuts import render
        project = self.get_project()
        if not can_manage_project(request.user, project):
            messages.error(request, "You do not have permission to import BoQ items for this project.")
            return redirect(reverse_lazy("budget:boq-list", kwargs={"project_pk": project.pk}))
        uploaded = request.FILES.get("file")
        if not uploaded:
            messages.error(request, "Please select a file to upload.")
            return render(request, self.template_name, {"project": project})

        ext = uploaded.name.rsplit(".", 1)[-1].lower()
        try:
            if ext == "csv":
                rows = self._parse_csv(uploaded)
            elif ext in ("xlsx", "xls"):
                rows = self._parse_excel(uploaded)
            else:
                messages.error(request, "Unsupported file type. Use CSV or Excel (.xlsx).")
                return render(request, self.template_name, {"project": project})
        except Exception as exc:
            messages.error(request, f"Error reading file: {exc}")
            return render(request, self.template_name, {"project": project})

        created = 0
        errors = []
        for i, row in enumerate(rows, start=2):  # start=2 (row 1 is header)
            try:
                item_number = str(row.get("item_number", "")).strip()
                description = str(row.get("description", "")).strip()
                if not item_number or not description:
                    errors.append(f"Row {i}: missing item_number or description — skipped.")
                    continue
                unit = str(row.get("unit", "nr")).strip().lower() or "nr"
                quantity = Decimal(str(row.get("quantity", "0") or "0").replace(",", ""))
                unit_rate = Decimal(str(row.get("unit_rate", "0") or "0").replace(",", ""))
                trade_section = str(row.get("trade_section", row.get("section", "OTHER")) or "OTHER").strip().upper()

                BoQItem.objects.update_or_create(
                    project=project,
                    item_number=item_number,
                    defaults={
                        "description": description,
                        "unit": unit,
                        "quantity": quantity,
                        "unit_rate": unit_rate,
                        "trade_section": trade_section,
                        "created_by": request.user,
                        "updated_by": request.user,
                    },
                )
                created += 1
            except Exception as exc:
                errors.append(f"Row {i}: {exc}")

        messages.success(request, f"Import complete: {created} items imported/updated.")
        if errors:
            for err in errors[:10]:
                messages.warning(request, err)
        return redirect(reverse_lazy("budget:boq-list", kwargs={"project_pk": project.pk}))

    def _parse_csv(self, f):
        import csv, io
        content = f.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        return [{k.strip().lower(): v for k, v in row.items()} for row in reader]

    def _parse_excel(self, f):
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip().lower() if h else f"col{i}" for i, h in enumerate(rows[0])]
        result = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            result.append({headers[i]: row[i] for i in range(len(headers))})
        return result

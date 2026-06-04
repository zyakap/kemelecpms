"""
Reusable Excel workbook builder utilities using openpyxl.
"""

import io
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# Brand colours (as ARGB hex for openpyxl)
KC_PRIMARY = "FF1A3A5C"  # dark navy
KC_ACCENT = "FF2980B9"   # blue
KC_LIGHT = "FFE8EDF3"    # pale blue-grey
KC_GREEN = "FF27AE60"
KC_AMBER = "FFE67E22"
KC_RED = "FFE74C3C"

THIN = Side(style="thin", color="FFCCCCCC")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header_fill():
    return PatternFill("solid", fgColor=KC_PRIMARY)


def _subheader_fill():
    return PatternFill("solid", fgColor=KC_LIGHT)


def _header_font(sz=11):
    return Font(bold=True, color="FFFFFFFF", size=sz, name="Arial")


def _title_font(sz=14):
    return Font(bold=True, color=KC_PRIMARY, size=sz, name="Arial")


def _bold():
    return Font(bold=True, name="Arial", size=10)


def _normal():
    return Font(name="Arial", size=10)


def make_workbook(title, subtitle=""):
    """Return a new Workbook with a cover sheet pre-formatted."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cover"

    ws.merge_cells("A1:G1")
    ws["A1"] = "KEMELE CONSTRUCTION"
    ws["A1"].font = Font(bold=True, size=16, color=KC_PRIMARY, name="Arial")

    ws.merge_cells("A2:G2")
    ws["A2"] = title
    ws["A2"].font = Font(bold=True, size=12, color=KC_ACCENT, name="Arial")

    if subtitle:
        ws.merge_cells("A3:G3")
        ws["A3"] = subtitle
        ws["A3"].font = Font(size=10, italic=True, name="Arial")

    ws.merge_cells("A4:G4")
    ws["A4"] = f"Generated: {date.today().strftime('%d %B %Y')}"
    ws["A4"].font = Font(size=9, color="FF888888", name="Arial")

    ws.row_dimensions[1].height = 22
    return wb


def add_sheet_header(ws, columns, title=None):
    """Add a header row to a worksheet. Returns the next available row."""
    start_row = 1
    if title:
        ws.merge_cells(f"A1:{get_column_letter(len(columns))}1")
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=12, color=KC_PRIMARY, name="Arial")
        ws["A1"].alignment = Alignment(horizontal="center")
        start_row = 2

    for col_idx, (header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.fill = _header_fill()
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    return start_row + 1


def write_row(ws, row_idx, values, bold=False, bg_color=None, number_format=None):
    """Write a row of values with optional styling."""
    fill = PatternFill("solid", fgColor=bg_color) if bg_color else None
    font = _bold() if bold else _normal()
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = font
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
        if number_format and isinstance(val, (int, float, Decimal)):
            cell.number_format = number_format
    return row_idx + 1


def workbook_response(wb, filename):
    """Return an HttpResponse with the workbook as an xlsx attachment."""
    from django.http import HttpResponse
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
